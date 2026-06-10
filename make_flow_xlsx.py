# -*- coding: utf-8 -*-
"""Generate the Now TrendIn 2.0 sign-in / membership design-flow workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Brand palette
GREEN  = "00C896"
ORANGE = "EE6A2A"
MAROON = "B5341B"
BLUE   = "2D7EEF"
GOLD   = "D4A017"
DARK   = "1A1A2E"
GREY   = "5B6472"
LBG    = "F4F5F7"
WHITE  = "FFFFFF"
BORDER = "E4E7EC"

thin = Side(style="thin", color=BORDER)
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

def fill(hexc):  return PatternFill("solid", fgColor=hexc)
def title_font(c=WHITE, sz=12): return Font(bold=True, color=c, size=sz, name="Calibri")
def body_font(c=DARK, sz=10, b=False): return Font(color=c, size=sz, bold=b, name="Calibri")

wb = openpyxl.Workbook()

# ───────────────────────── Sheet 1: Sign-in Flow ─────────────────────────
ws = wb.active
ws.title = "Sign-in Flow"
ws.sheet_view.showGridLines = False
widths = [3, 26, 30, 34, 30, 3]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 2
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws.cell(r, 2, "Now TrendIn 2.0  —  Sign-in & Onboarding Flow")
c.font = Font(bold=True, color=DARK, size=16, name="Calibri")
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws.cell(r, 2, "Current build (Expo / React Native + Django JWT backend).  Status legend: ✅ Live  ·  ⚠️ Built, needs config  ·  ❌ Not built")
c.font = Font(italic=True, color=GREY, size=10)
r += 2

# Flow steps: (step, screen/route, what happens, status)
header = ["Step", "Screen / Route", "What happens", "Status"]
for j, h in enumerate(header, 2):
    cell = ws.cell(r, j, h)
    cell.fill = fill(DARK); cell.font = title_font(); cell.border = box
    cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 22
r += 1

steps = [
    ("1", "Splash  /splash", "Animated logo (2s). Checks stored JWT → if valid, hydrates user and skips ahead. Else routes to onboarding (first run) or login.", "✅ Live"),
    ("2", "Onboarding  /onboarding", "First-run intro slides. Sets 'onboarding_seen' flag so it shows only once.", "✅ Live"),
    ("3", "Login  /login", "Email + password → Django /api/auth/login/ → JWT. Routes to app (if tier set) or membership.", "✅ Live"),
    ("3a", "Login → Continue with Google", "Google OAuth (expo-auth-session ID-token flow). Web + iOS client IDs live; backend verifies token at /api/auth/google/ and issues JWT.", "✅ Live (Android client deferred)"),
    ("3b", "Login → Forgot password?", "Links to reset flow (see step 6).", "✅ Live"),
    ("4", "Sign up  /signup", "Name + email + password + confirm. Zod rules (8+ char, uppercase, number). T&C checkbox blocks submit. → /api/auth/signup/ → JWT → membership.", "✅ Live"),
    ("5", "Membership  /membership", "Choose Consumer / Business / Enterprise. → PATCH /api/auth/me/ saves tier → enters app.", "✅ Live"),
    ("6", "Forgot password  /forgot-password", "3 steps: (a) enter email → 6-digit code emailed; (b) enter code + new password; (c) done. Code expires 15 min. Email delivery via Gmail SMTP (no-reply@joinmynet.com) — tested & live.", "✅ Live"),
    ("7", "App  /(app)", "Tier-aware tab bar: Trends · Search · Grade · Alerts · Profile. Data gated by tier freshness window.", "✅ Live"),
]
for s in steps:
    statuscol = s[3]
    sc = GREEN if statuscol.startswith("✅") else (GOLD if statuscol.startswith("⚠️") else MAROON)
    cells = [
        (s[0], "center", DARK, False, WHITE),
        (s[1], "left", DARK, True, WHITE),
        (s[2], "left", GREY, False, WHITE),
        (s[3], "left", sc, True, WHITE),
    ]
    for j, (val, al, color, bold, bg) in enumerate(cells, 2):
        cell = ws.cell(r, j, val)
        cell.font = body_font(color, 10, bold); cell.border = box
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal=al, vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 46
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws.cell(r, 2, "Backend: Django REST + SimpleJWT on Heroku (app: nowtrendin-backend).  Scoring engine: FastAPI (app: nowtrendin).")
c.font = Font(italic=True, color=GREY, size=9)

# ───────────────────────── Sheet 2: Membership Matrix ─────────────────────
ws2 = wb.create_sheet("Membership Tiers")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 3
ws2.column_dimensions["B"].width = 34
for col in ("C", "D", "E"):
    ws2.column_dimensions[col].width = 26

r = 2
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws2.cell(r, 2, "Membership Tiers  —  Feature & Access Matrix")
c.font = Font(bold=True, color=DARK, size=16)
r += 2

# Header row with tier names
tiers = [("Consumer", BLUE), ("Business", GREEN), ("Enterprise", GOLD)]
ws2.cell(r, 2, "Feature").fill = fill(DARK)
ws2.cell(r, 2).font = title_font(); ws2.cell(r, 2).border = box
for j, (name, color) in enumerate(tiers, 3):
    cell = ws2.cell(r, j, name)
    cell.fill = fill(color); cell.font = title_font(); cell.border = box
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[r].height = 24
r += 1

rows = [
    ("Price / month", "$49", "$499", "$250,000"),
    ("Seats", "1", "1", "5 (shared token pool)"),
    ("Data freshness", "24h+ old only", "12h+ old only", "Live (the moment data is obtained)"),
    ("Gradient Score history", "✓", "✓", "✓"),
    ("Trend monitoring feed", "✓", "✓", "✓"),
    ("Email + push alerts", "✓", "✓", "✓ (custom thresholds)"),
    ("Full signal search + filter", "✗", "✓", "✓"),
    ("Query NEW trend topics", "✗", "✗", "✓ (1 token / search)"),
    ("Direct topic query", "✗", "✗", "✓"),
    ("AI-grade unknown topic", "✗", "✗", "✓ (1 token)"),
    ("Edit data-source weights", "✗", "✗", "✓"),
    ("API access", "✗", "✗", "✓"),
    ("Query tokens / month", "0", "0", "100,000 (shared)"),
]
for row in rows:
    label = ws2.cell(r, 2, row[0])
    label.font = body_font(DARK, 10, True); label.border = box; label.fill = fill(WHITE)
    label.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for j, val in enumerate(row[1:], 3):
        cell = ws2.cell(r, j, val)
        color = MAROON if val == "✗" else (GREEN if val == "✓" else DARK)
        cell.font = body_font(color, 10, val in ("✓", "✗"))
        cell.border = box; cell.fill = fill(WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[r].height = 20
    r += 1

r += 1
ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws2.cell(r, 2, "Access rules enforced centrally in constants/tiers.ts via canAccess() and isDataAccessible(). "
                   "Data-aging waterfall: a new score is Enterprise-only (live) → visible to Business at 12h → Consumer at 24h.")
c.font = Font(italic=True, color=GREY, size=9)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws2.row_dimensions[r].height = 40

# ───────────────────────── Sheet 3: To-Fix / Backlog ──────────────────────
ws3 = wb.create_sheet("What Needs Fixing")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 3
ws3.column_dimensions["B"].width = 26
ws3.column_dimensions["C"].width = 40
ws3.column_dimensions["D"].width = 30
ws3.column_dimensions["E"].width = 14

r = 2
ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
c = ws3.cell(r, 2, "Sign-in Flow  —  What Still Needs Fixing")
c.font = Font(bold=True, color=DARK, size=16)
r += 2

hdr = ["Item", "Current state", "What's needed to finish", "Priority"]
for j, h in enumerate(hdr, 2):
    cell = ws3.cell(r, j, h)
    cell.fill = fill(DARK); cell.font = title_font(); cell.border = box
ws3.row_dimensions[r].height = 22
r += 1

backlog = [
    ("Google sign-in", "DONE for Web + iOS. expo-auth-session ID-token flow; backend /api/auth/google/ verifies token + issues JWT. Works in Expo Go and on web.",
     "Android client deferred: needs SHA-1 from first native build. After build, create an Android OAuth client (package com.nowtrendin.app + SHA-1), paste into constants/google.ts + set GOOGLE_ANDROID_CLIENT_ID on backend.", "DONE"),
    ("Forgot-password email", "DONE. Full 3-step UI + backend; Gmail SMTP (no-reply@joinmynet.com) configured on nowtrendin-backend. Live test email delivered successfully.",
     "None. (If volume grows past Gmail's ~500/day, move to SendGrid/Mailgun — no code change.)", "DONE"),
    ("Profile notify prefs typecheck", "DONE. notifyEmail/notifyPush added to the User type; frontend typechecks 100% clean.",
     "None.", "DONE"),
    ("Native app identifiers", "DONE. iOS bundleIdentifier + Android package set to com.nowtrendin.app in app.json.",
     "Confirm the bundle ID before first store submission — it is effectively permanent once published.", "DONE"),
    ("First native build (EAS)", "Not yet built. Required to unlock Android Google client + Apple sign-in.",
     "Run `eas build -p ios` / `eas build -p android`. Then `eas credentials` to read the Android SHA-1.", "MED (native launch)"),
    ("Apple sign-in", "Not present.",
     "iOS App Store requires Apple sign-in when Google sign-in ships. Add Sign in with Apple alongside Google before submission.", "MED (iOS launch)"),
    ("SMS 2FA (phone)", "Backend send/verify built; requires Twilio creds. Not wired into the sign-up flow yet.",
     "Set Twilio creds if 2FA is desired; optionally add to onboarding.", "LOW"),
]
for item in backlog:
    pr = item[3]
    pc = GREEN if pr == "DONE" else (MAROON if pr == "HIGH" else (GOLD if pr.startswith("MED") else GREY))
    vals = [
        (item[0], "left", DARK, True),
        (item[1], "left", GREY, False),
        (item[2], "left", GREY, False),
        (item[3], "center", pc, True),
    ]
    for j, (val, al, color, bold) in enumerate(vals, 2):
        cell = ws3.cell(r, j, val)
        cell.font = body_font(color, 10, bold); cell.border = box; cell.fill = fill(WHITE)
        cell.alignment = Alignment(horizontal=al, vertical="top", wrap_text=True)
    ws3.row_dimensions[r].height = 60
    r += 1

out = r"C:\Users\acinv\OneDrive\Desktop\CODING PROJECTS\NowTrendin v2.0\NowTrendin_SignIn_Flow.xlsx"
wb.save(out)
print("Saved:", out)
